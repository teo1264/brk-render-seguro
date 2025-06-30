#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gerador Excel BRK - SIMPLES E DIRETO
Pega dados prontos da faturas_brk + gera Excel formatado
"""

import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict
from datetime import datetime
import os
import io
import requests
from flask import jsonify, request, send_file
import logging

# Imports do sistema existente
from auth.microsoft_auth import MicrosoftAuth

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelGeneratorBRK:
    """Gerador Excel BRK - pega dados prontos da faturas_brk"""
    
    def __init__(self):
        self.mes_nomes = {
            1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
            5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto", 
            9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
        }
    
    def handle_request(self):
        """Handle HTTP request para geração Excel"""
        try:
            # Verificar autenticação Microsoft
            auth_result = self._verificar_autenticacao()
            if auth_result:
                return auth_result
            
            if request.method == 'GET':
                return self._render_form()
            elif request.method == 'POST':
                return self._processar_geracao()
        except Exception as e:
            logger.error(f"Erro handle_request: {e}")
            return jsonify({"erro": str(e)}), 500
    
    def _verificar_autenticacao(self):
        """Verificar autenticação Microsoft"""
        try:
            self.auth = MicrosoftAuth()
            
            if not self.auth.access_token:
                return jsonify({
                    "erro": "Token Microsoft não encontrado",
                    "action": "redirect",
                    "url": "/upload-token"
                }), 401
            
            if not self.auth.validar_token():
                logger.info("Token expirado, tentando renovar...")
                if not self.auth.atualizar_token():
                    return jsonify({
                        "erro": "Token Microsoft expirado - faça login novamente",
                        "action": "redirect", 
                        "url": "/upload-token"
                    }), 401
            
            logger.info("✅ Autenticação Microsoft validada")
            return None  # Auth OK
            
        except Exception as e:
            logger.error(f"Erro verificação auth: {e}")
            return jsonify({
                "erro": "Erro de autenticação Microsoft",
                "details": str(e)
            }), 500
    
    def _render_form(self):
        """Renderizar formulário simples"""
        hoje = datetime.now()
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Gerador Planilha BRK</title>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial; margin: 40px; background: #f5f5f5; }}
                .container {{ background: white; padding: 30px; border-radius: 10px; max-width: 500px; }}
                h2 {{ color: #2c5282; margin-bottom: 20px; }}
                .form-group {{ margin-bottom: 15px; }}
                label {{ display: block; margin-bottom: 5px; font-weight: bold; }}
                select, button {{ padding: 10px; border: 1px solid #ddd; border-radius: 5px; width: 100%; }}
                button {{ background: #2c5282; color: white; font-weight: bold; cursor: pointer; }}
                button:hover {{ background: #2a4e7a; }}
                .info {{ background: #e6f3ff; padding: 15px; border-radius: 5px; margin-top: 20px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h2>📊 Gerador Planilha BRK</h2>
                <form method="POST">
                    <div class="form-group">
                        <label>Mês:</label>
                        <select name="mes" required>
                            <option value="">Selecione o mês</option>
                            <option value="1">Janeiro</option>
                            <option value="2">Fevereiro</option>
                            <option value="3">Março</option>
                            <option value="4">Abril</option>
                            <option value="5">Maio</option>
                            <option value="6" {"selected" if hoje.month == 6 else ""}>Junho</option>
                            <option value="7" {"selected" if hoje.month == 7 else ""}>Julho</option>
                            <option value="8">Agosto</option>
                            <option value="9">Setembro</option>
                            <option value="10">Outubro</option>
                            <option value="11">Novembro</option>
                            <option value="12">Dezembro</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Ano:</label>
                        <select name="ano" required>
                            <option value="">Selecione o ano</option>
                            <option value="2024">2024</option>
                            <option value="2025" selected>2025</option>
                            <option value="2026">2026</option>
                        </select>
                    </div>
                    <button type="submit">🚀 Gerar Planilha Excel</button>
                </form>
                
                <div class="info">
                    <strong>ℹ️ Sistema:</strong><br>
                    • Dados da tabela faturas_brk (já processados)<br>
                    • PIA separada das demais casas<br>
                    • Casas faltantes detectadas automaticamente<br>
                    • Planilha será salva no OneDrive /BRK/Faturas/<br>
                    • Seção de controle para duplicatas (se houver)
                </div>
            </div>
        </body>
        </html>
        """
        return html
    
    def _processar_geracao(self):
        """Processar solicitação de geração Excel"""
        try:
            # Validar parâmetros
            mes, ano = self._validar_parametros()
            
            logger.info(f"Gerando planilha BRK para {self.mes_nomes[mes]}/{ano}")
            
            # Gerar planilha
            excel_bytes = self.gerar_planilha_mensal(mes, ano)
            
            # Nome do arquivo
            nome_arquivo = f"BRK-Planilha-{ano}-{mes:02d}.xlsx"
            
            # SALVAR APENAS NO ONEDRIVE (sem download)
            try:
                self._salvar_onedrive_background(excel_bytes, mes, ano)
                
                # Página de sucesso
                return self._render_sucesso(mes, ano, nome_arquivo)
                
            except Exception as e:
                logger.error(f"Erro salvando OneDrive: {e}")
                return jsonify({"erro": f"Erro salvando no OneDrive: {e}"}), 500
            
        except Exception as e:
            logger.error(f"Erro _processar_geracao: {e}")
            return jsonify({"erro": str(e)}), 500
    
    def _validar_parametros(self):
        """Validar parâmetros (segurança)"""
        try:
            mes = int(request.form.get('mes', 0))
            ano = int(request.form.get('ano', 0))
            
            if not (1 <= mes <= 12):
                raise ValueError("Mês deve estar entre 1 e 12")
            
            if not (2020 <= ano <= 2030):
                raise ValueError("Ano deve estar entre 2020 e 2030")
            
            return mes, ano
            
        except (ValueError, TypeError) as e:
            raise ValueError(f"Parâmetros inválidos: {e}")
    
    def _render_sucesso(self, mes, ano, nome_arquivo):
        """Página de sucesso após salvar no OneDrive"""
        pasta_onedrive = f"/BRK/Faturas/{ano}/{mes:02d}/"
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>✅ Planilha BRK Gerada</title>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial; margin: 40px; background: #f5f5f5; }}
                .container {{ background: white; padding: 30px; border-radius: 10px; max-width: 600px; margin: 0 auto; }}
                .success {{ color: #28a745; font-weight: bold; font-size: 18px; }}
                .info {{ background: #e6f3ff; padding: 15px; border-radius: 5px; margin: 15px 0; }}
                .button {{ background: #2c5282; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; margin: 10px 5px; display: inline-block; }}
                .button:hover {{ background: #2a4e7a; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h2>✅ Planilha BRK Gerada com Sucesso!</h2>
                
                <div class="success">
                    📊 Relatório de {self.mes_nomes[mes]}/{ano} processado
                </div>
                
                <div class="info">
                    <strong>📁 Arquivo salvo no OneDrive:</strong><br>
                    <code>{pasta_onedrive}{nome_arquivo}</code><br><br>
                    
                    <strong>📋 Conteúdo da planilha:</strong><br>
                    • PIA separada (Conta Bancária A)<br>
                    • Casas de Oração agrupadas por vencimento (Conta Bancária B)<br>
                    • Casas faltantes detectadas automaticamente<br>
                    • Totais e subtotais calculados<br>
                    • Seção de controle para duplicatas (se houver)
                </div>
                
                <div style="text-align: center; margin-top: 30px;">
                    <a href="/gerar-planilha-brk" class="button">📊 Gerar Outra Planilha</a>
                    <a href="/" class="button">🏠 Dashboard</a>
                </div>
            </div>
        </body>
        </html>
        """
        return html
    
    def gerar_planilha_mensal(self, mes, ano):
        """MÉTODO PRINCIPAL: Gerar planilha Excel do período"""
        try:
            logger.info(f"Iniciando geração Excel {mes}/{ano}")
            
            # 1. BUSCAR DADOS NORMAIS da faturas_brk
            faturas_normais = self._buscar_faturas_prontas(mes, ano)
            logger.info(f"Faturas NORMAIS: {len(faturas_normais)}")
            
            # 2. CARREGAR BASE COMPLETA OneDrive (para casas faltantes)
            base_completa = self._carregar_base_onedrive()
            logger.info(f"Casas na base OneDrive: {len(base_completa)}")
            
            # 3. DETECTAR CASAS FALTANTES
            casas_faltantes = self._detectar_casas_faltantes(faturas_normais, base_completa, mes, ano)
            logger.info(f"Casas faltantes: {len(casas_faltantes)}")
            
            # 4. COMBINAR: faturas normais + faltantes
            dados_completos = faturas_normais + casas_faltantes
            
            # 5. SEPARAR PIA das demais casas
            dados_pia, dados_casas = self._separar_pia_casas(dados_completos)
            logger.info(f"PIAs: {len(dados_pia)}, Casas: {len(dados_casas)}")
            
            # 6. BUSCAR OUTROS STATUS (simples, separado)
            try:
                faturas_outros = self._buscar_outros_status_simples(mes, ano)
                logger.info(f"Outros status: {len(faturas_outros)}")
            except:
                faturas_outros = []
            
            # 7. GERAR EXCEL FORMATADO
            excel_bytes = self._gerar_excel_com_controle(dados_pia, dados_casas, faturas_outros, mes, ano)
            
            logger.info(f"Planilha gerada: {len(excel_bytes)} bytes")
            return excel_bytes
            
        except Exception as e:
            logger.error(f"Erro gerar_planilha_mensal: {e}")
            raise
    
    def _buscar_faturas_prontas(self, mes, ano):
        """BUSCAR DADOS NORMAIS da tabela faturas_brk"""
        try:
            # Usar DatabaseBRK existente (sistema já configurado)
            from processor.database_brk import DatabaseBRK
            
            onedrive_brk_id = os.getenv("ONEDRIVE_BRK_ID")
            if not onedrive_brk_id:
                raise ValueError("ONEDRIVE_BRK_ID não configurado")
            
            # DatabaseBRK já funciona
            db = DatabaseBRK(self.auth, onedrive_brk_id)
            conn = db.get_connection()
            
            if not conn:
                raise ValueError("Conexão database não disponível")
            
            conn.row_factory = sqlite3.Row
            
            # Query SIMPLES nos dados NORMAIS (como era antes)
            query = """
                SELECT * FROM faturas_brk 
                WHERE competencia LIKE ? 
                AND vencimento LIKE ?
                AND status_duplicata = 'NORMAL'
                ORDER BY vencimento, casa_oracao
            """
            
            params = (f"%/{ano}", f"__/{mes:02d}/%")
            
            cursor = conn.execute(query, params)
            resultados = cursor.fetchall()
            
            # Converter para lista de dicts
            faturas = []
            for row in resultados:
                fatura = dict(row)
                faturas.append(fatura)
            
            logger.info(f"✅ Dados NORMAIS: {len(faturas)} registros")
            return faturas
            
        except Exception as e:
            logger.error(f"Erro _buscar_faturas_prontas: {e}")
            raise
    
    def _buscar_outros_status_simples(self, mes, ano):
        """Buscar apenas outros status (DUPLICATA, CUIDADO, etc.) - SIMPLES"""
        try:
            from processor.database_brk import DatabaseBRK
            
            onedrive_brk_id = os.getenv("ONEDRIVE_BRK_ID")
            db = DatabaseBRK(self.auth, onedrive_brk_id)
            conn = db.get_connection()
            
            if not conn:
                return []
            
            conn.row_factory = sqlite3.Row
            
            # Query SIMPLES para outros status
            query = """
                SELECT * FROM faturas_brk 
                WHERE competencia LIKE ? 
                AND vencimento LIKE ?
                AND status_duplicata != 'NORMAL'
                ORDER BY status_duplicata, casa_oracao
            """
            
            params = (f"%/{ano}", f"__/{mes:02d}/%")
            
            cursor = conn.execute(query, params)
            resultados = cursor.fetchall()
            
            # Converter para lista de dicts
            outros = []
            for row in resultados:
                outro = dict(row)
                outros.append(outro)
            
            return outros
            
        except Exception as e:
            logger.warning(f"Erro buscando outros status: {e}")
            return []
    
    def _carregar_base_onedrive(self):
        """Carregar CDC_BRK_CCB.xlsx do OneDrive"""
        try:
            logger.info("Carregando CDC_BRK_CCB.xlsx do OneDrive...")
            
            headers = self.auth.obter_headers_autenticados()
            url = "https://graph.microsoft.com/v1.0/me/drive/root:/BRK/CDC_BRK_CCB.xlsx:/content"
            
            response = requests.get(url, headers=headers, timeout=30)
            
            if response.status_code == 200:
                logger.info("✅ CDC_BRK_CCB.xlsx baixado do OneDrive")
                return self._processar_excel_base(response.content)
                
            elif response.status_code == 401:
                if self.auth.tentar_renovar_se_necessario(401):
                    headers = self.auth.obter_headers_autenticados()
                    response = requests.get(url, headers=headers, timeout=30)
                    if response.status_code == 200:
                        return self._processar_excel_base(response.content)
                
                logger.error("Token Microsoft expirado")
                return []
                
            elif response.status_code == 404:
                logger.error("CDC_BRK_CCB.xlsx não encontrado no OneDrive /BRK/")
                return []
                
            else:
                logger.error(f"Erro baixando CDC_BRK_CCB.xlsx: HTTP {response.status_code}")
                return []
            
        except Exception as e:
            logger.error(f"Erro _carregar_base_onedrive: {e}")
            return []
    
    def _processar_excel_base(self, excel_content):
        """Processar Excel CDC_BRK_CCB.xlsx"""
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(excel_content))
            worksheet = workbook.active
            
            base_completa = []
            
            # Processar linhas (skip header)
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:  # Tem Casa e CDC
                    casa = str(row[0]).strip()
                    cdc = str(row[1]).strip()
                    dia_vencimento = row[4] if len(row) > 4 and row[4] else 1
                    
                    try:
                        dia_vencimento = int(dia_vencimento)
                    except:
                        dia_vencimento = 1
                    
                    base_completa.append({
                        "cdc": cdc,
                        "casa": casa,
                        "dia_vencimento": dia_vencimento
                    })
            
            logger.info(f"✅ Base OneDrive processada: {len(base_completa)} casas")
            return base_completa
            
        except Exception as e:
            logger.error(f"Erro processando Excel base: {e}")
            return []
    
    def _detectar_casas_faltantes(self, faturas_prontas, base_completa, mes, ano):
        """Detectar casas que não receberam fatura"""
        try:
            # CDCs que já estão na base (chegaram por email)
            cdcs_processados = {fatura["cdc"] for fatura in faturas_prontas}
            
            # CDCs da base completa
            cdcs_base = {casa["cdc"] for casa in base_completa}
            
            # CDCs faltantes
            cdcs_faltantes = cdcs_base - cdcs_processados
            
            casas_faltantes = []
            mes_nome = self.mes_nomes[mes]
            
            for casa in base_completa:
                if casa["cdc"] in cdcs_faltantes:
                    dia_venc = casa["dia_vencimento"]
                    vencimento = f"{dia_venc:02d}/{mes:02d}/{ano}"
                    
                    casa_faltante = {
                        "cdc": casa["cdc"],
                        "casa_oracao": casa["casa"],
                        "competencia": f"{mes_nome}/{ano}",
                        "vencimento": vencimento,
                        "valor": "",
                        "nota_fiscal": "",
                        "data_emissao": "",
                        "medido_real": None,
                        "faturado": None,
                        "media_6m": None,
                        "porcentagem_consumo": "",
                        "alerta_consumo": "Não recebido",
                        "status_duplicata": "FALTANTE"
                    }
                    
                    casas_faltantes.append(casa_faltante)
            
            return casas_faltantes
            
        except Exception as e:
            logger.error(f"Erro _detectar_casas_faltantes: {e}")
            return []
    
    def _separar_pia_casas(self, dados_completos):
        """Separar PIA das demais casas"""
        dados_pia = []
        dados_casas = []
        
        for registro in dados_completos:
            if self._eh_pia(registro):
                dados_pia.append(registro)
            else:
                dados_casas.append(registro)
        
        return dados_pia, dados_casas
    
    def _eh_pia(self, registro):
        """Identificar se é PIA"""
        casa = registro.get("casa_oracao", "").upper()
        return casa == "PIA"
    
    def _gerar_excel_com_controle(self, dados_pia, dados_casas, faturas_outros, mes, ano):
        """Gerar Excel formatado com seção de controle no final"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"BRK {self.mes_nomes[mes]} {ano}"
            
            linha_atual = 1
            
            # SEÇÃO PRINCIPAL: APENAS NORMAIS
            # Título principal
            linha_atual = self._adicionar_titulo_principal(ws, linha_atual, mes, ano)
            
            # Seção PIA (NORMAIS)
            linha_atual = self._adicionar_secao_pia(ws, linha_atual, dados_pia)
            
            # Seção Casas agrupadas por vencimento (NORMAIS)
            linha_atual = self._adicionar_secao_casas(ws, linha_atual, dados_casas)
            
            # Totais finais (NORMAIS)
            linha_atual = self._adicionar_totais_finais(ws, linha_atual, dados_pia, dados_casas)
            
            # SEÇÃO ADICIONAL: OUTROS STATUS (se existirem)
            if faturas_outros:
                linha_atual = self._adicionar_secao_controle_simples(ws, linha_atual + 3, faturas_outros)
            
            # Formatação geral
            self._aplicar_formatacao_geral(ws)
            
            # Salvar em bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.read()
            
        except Exception as e:
            logger.error(f"Erro _gerar_excel_com_controle: {e}")
            raise
    
    def _adicionar_titulo_principal(self, ws, linha, mes, ano):
        """Título principal"""
        titulo = f"📋 RELATÓRIO BRK - {self.mes_nomes[mes].upper()}/{ano}"
        
        ws.merge_cells(f"A{linha}:L{linha}")  # 12 colunas A-L
        ws[f"A{linha}"] = titulo
        ws[f"A{linha}"].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f"A{linha}"].fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
        ws[f"A{linha}"].alignment = Alignment(horizontal="center", vertical="center")
        
        return linha + 2
    
    def _adicionar_secao_pia(self, ws, linha_inicial, dados_pia):
        """Seção PIA"""
        linha = linha_inicial
        
        # Cabeçalho PIA (12 colunas)
        ws.merge_cells(f"A{linha}:L{linha}")
        ws[f"A{linha}"] = "=== PIA (Conta Bancária A) ==="
        ws[f"A{linha}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{linha}"].fill = PatternFill(start_color="C53030", end_color="C53030", fill_type="solid")
        ws[f"A{linha}"].alignment = Alignment(horizontal="center")
        linha += 1
        
        # Headers COMPLETOS (igual ao relatório original)
        headers = ["CDC", "Casa de Oração", "Competência", "Data Emissão", "Vencimento", "Nota Fiscal", "Valor", "Medido Real", "Faturado", "Média 6M", "% Consumo", "Alerta Consumo"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=linha, column=col, value=header)
            ws.cell(row=linha, column=col).font = Font(bold=True)
        linha += 1
        
        # Dados PIA COMPLETOS
        subtotal_pia = 0
        for pia in dados_pia:
            ws[f"A{linha}"] = pia.get("cdc", "")
            ws[f"B{linha}"] = pia.get("casa_oracao", "")
            ws[f"C{linha}"] = pia.get("competencia", "")
            ws[f"D{linha}"] = pia.get("data_emissao", "")
            ws[f"E{linha}"] = pia.get("vencimento", "")
            ws[f"F{linha}"] = pia.get("nota_fiscal", "")
            ws[f"G{linha}"] = pia.get("valor", "")
            ws[f"H{linha}"] = pia.get("medido_real", "")
            ws[f"I{linha}"] = pia.get("faturado", "")
            ws[f"J{linha}"] = pia.get("media_6m", "")
            ws[f"K{linha}"] = pia.get("porcentagem_consumo", "")
            ws[f"L{linha}"] = pia.get("alerta_consumo", "")
            
            # Somar valor se numérico (coluna G agora)
            if pia.get("valor"):
                try:
                    valor_num = float(pia["valor"].replace("R$", "").replace(",", ".").strip())
                    subtotal_pia += valor_num
                except:
                    pass
            
            linha += 1
        
        # Subtotal PIA (ajustar para coluna G)
        ws.merge_cells(f"A{linha}:F{linha}")
        ws[f"A{linha}"] = "SUBTOTAL PIA:"
        ws[f"A{linha}"].font = Font(bold=True)
        ws[f"G{linha}"] = f"R$ {subtotal_pia:.2f}".replace(".", ",")
        ws[f"G{linha}"].font = Font(bold=True)
        
        return linha + 2
    
    def _adicionar_secao_casas(self, ws, linha_inicial, dados_casas):
        """Seção Casas agrupadas por vencimento"""
        linha = linha_inicial
        
        # Cabeçalho Casas (12 colunas)
        ws.merge_cells(f"A{linha}:L{linha}")
        ws[f"A{linha}"] = "=== CASAS DE ORAÇÃO (Conta Bancária B) ==="
        ws[f"A{linha}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{linha}"].fill = PatternFill(start_color="2D7D32", end_color="2D7D32", fill_type="solid")
        ws[f"A{linha}"].alignment = Alignment(horizontal="center")
        linha += 1
        
        # Agrupar por vencimento
        casas_por_vencimento = defaultdict(list)
        for casa in dados_casas:
            vencimento = casa.get("vencimento", "")
            casas_por_vencimento[vencimento].append(casa)
        
        subtotal_casas = 0
        
        # Para cada vencimento
        for vencimento in sorted(casas_por_vencimento.keys()):
            if not vencimento:
                continue
                
            # Cabeçalho do vencimento (12 colunas)
            ws.merge_cells(f"A{linha}:L{linha}")
            ws[f"A{linha}"] = f"Vencimento {vencimento}:"
            ws[f"A{linha}"].font = Font(bold=True, color="2D7D32")
            linha += 1
            
            # Headers COMPLETOS
            headers = ["CDC", "Casa de Oração", "Competência", "Data Emissão", "Vencimento", "Nota Fiscal", "Valor", "Medido Real", "Faturado", "Média 6M", "% Consumo", "Alerta Consumo"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=linha, column=col, value=header)
                ws.cell(row=linha, column=col).font = Font(bold=True, size=9)
            linha += 1
            
            subtotal_vencimento = 0
            
            # Casas do vencimento COM TODOS OS CAMPOS
            for casa in casas_por_vencimento[vencimento]:
                ws[f"A{linha}"] = casa.get("cdc", "")
                ws[f"B{linha}"] = casa.get("casa_oracao", "")
                ws[f"C{linha}"] = casa.get("competencia", "")
                ws[f"D{linha}"] = casa.get("data_emissao", "")
                ws[f"E{linha}"] = casa.get("vencimento", "")
                ws[f"F{linha}"] = casa.get("nota_fiscal", "")
                ws[f"G{linha}"] = casa.get("valor", "")
                ws[f"H{linha}"] = casa.get("medido_real", "")
                ws[f"I{linha}"] = casa.get("faturado", "")
                ws[f"J{linha}"] = casa.get("media_6m", "")
                ws[f"K{linha}"] = casa.get("porcentagem_consumo", "")
                ws[f"L{linha}"] = casa.get("alerta_consumo", "")
                
                # Somar valor (coluna G agora)
                if casa.get("valor"):
                    try:
                        valor_num = float(casa["valor"].replace("R$", "").replace(",", ".").strip())
                        subtotal_vencimento += valor_num
                        subtotal_casas += valor_num
                    except:
                        pass
                
                linha += 1
            
            # Subtotal do vencimento (ajustar para coluna G)
            ws.merge_cells(f"A{linha}:F{linha}")
            ws[f"A{linha}"] = f"SUBTOTAL {vencimento}:"
            ws[f"A{linha}"].font = Font(bold=True, size=9)
            ws[f"G{linha}"] = f"R$ {subtotal_vencimento:.2f}".replace(".", ",")
            ws[f"G{linha}"].font = Font(bold=True, size=9)
            linha += 2
        
        # Subtotal todas as casas (ajustar para coluna G)
        ws.merge_cells(f"A{linha}:F{linha}")
        ws[f"A{linha}"] = "SUBTOTAL CASAS:"
        ws[f"A{linha}"].font = Font(bold=True)
        ws[f"G{linha}"] = f"R$ {subtotal_casas:.2f}".replace(".", ",")
        ws[f"G{linha}"].font = Font(bold=True)
        
        return linha + 1
    
    def _adicionar_totais_finais(self, ws, linha, dados_pia, dados_casas):
        """Totais finais"""
        total_geral = 0
        
        # Somar PIAs
        for pia in dados_pia:
            if pia.get("valor"):
                try:
                    valor_num = float(pia["valor"].replace("R$", "").replace(",", ".").strip())
                    total_geral += valor_num
                except:
                    pass
        
        # Somar Casas
        for casa in dados_casas:
            if casa.get("valor"):
                try:
                    valor_num = float(casa["valor"].replace("R$", "").replace(",", ".").strip())
                    total_geral += valor_num
                except:
                    pass
        
        # Total geral (ajustar para coluna G)
        ws.merge_cells(f"A{linha}:F{linha}")
        ws[f"A{linha}"] = "TOTAL GERAL:"
        ws[f"A{linha}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"A{linha}"].fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        ws[f"G{linha}"] = f"R$ {total_geral:.2f}".replace(".", ",")
        ws[f"G{linha}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"G{linha}"].fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        
        return linha + 1
    
    def _aplicar_formatacao_geral(self, ws):
        """Formatação geral com todas as colunas"""
        # Largura colunas (12 colunas A-L)
        ws.column_dimensions['A'].width = 12  # CDC
        ws.column_dimensions['B'].width = 35  # Casa de Oração
        ws.column_dimensions['C'].width = 15  # Competência
        ws.column_dimensions['D'].width = 12  # Data Emissão
        ws.column_dimensions['E'].width = 12  # Vencimento
        ws.column_dimensions['F'].width = 15  # Nota Fiscal
        ws.column_dimensions['G'].width = 15  # Valor
        ws.column_dimensions['H'].width = 12  # Medido Real
        ws.column_dimensions['I'].width = 12  # Faturado
        ws.column_dimensions['J'].width = 12  # Média 6M
        ws.column_dimensions['K'].width = 15  # % Consumo
        ws.column_dimensions['L'].width = 20  # Alerta Consumo
        
        # Bordas para todas as células com dados
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = thin_border
    
    def _adicionar_secao_controle_simples(self, ws, linha_inicial, faturas_outros):
        """Seção simples de controle para outros status"""
        linha = linha_inicial
        
        # Cabeçalho simples
        ws.merge_cells(f"A{linha}:L{linha}")
        ws[f"A{linha}"] = f"=== CONTROLE: {len(faturas_outros)} FATURAS COM STATUS ESPECIAL ==="
        ws[f"A{linha}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{linha}"].fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
        ws[f"A{linha}"].alignment = Alignment(horizontal="center")
        linha += 1
        
        # Explicação
        ws.merge_cells(f"A{linha}:L{linha}")
        ws[f"A{linha}"] = "⚠️ Faturas abaixo NÃO estão incluídas nos totais acima"
        ws[f"A{linha}"].font = Font(bold=True, italic=True, color="FF6600")
        ws[f"A{linha}"].alignment = Alignment(horizontal="center")
        linha += 2
        
        # Headers simples
        headers = ["CDC", "Casa de Oração", "Competência", "Vencimento", "Valor", "Status", "Observação"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=linha, column=col, value=header)
            ws.cell(row=linha, column=col).font = Font(bold=True, size=9)
            ws.cell(row=linha, column=col).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        linha += 1
        
        # Dados simples
        for fatura in faturas_outros:
            ws[f"A{linha}"] = fatura.get("cdc", "")
            ws[f"B{linha}"] = fatura.get("casa_oracao", "")
            ws[f"C{linha}"] = fatura.get("competencia", "")
            ws[f"D{linha}"] = fatura.get("vencimento", "")
            ws[f"E{linha}"] = fatura.get("valor", "")
            ws[f"F{linha}"] = fatura.get("status_duplicata", "")
            ws[f"G{linha}"] = "Verificar manualmente"
            
            # Cor de fundo leve
            for col in range(1, 8):
                cell = ws.cell(row=linha, column=col)
                cell.fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
            
            linha += 1
        
        return linha
    
    def _salvar_onedrive_background(self, excel_bytes, mes, ano):
        """Salvar no OneDrive de forma síncrona"""
        try:
            nome_arquivo = f"BRK-Planilha-{ano}-{mes:02d}.xlsx"
            pasta_destino = f"/BRK/Faturas/{ano}/{mes:02d}/"
            
            logger.info(f"Salvando no OneDrive: {nome_arquivo}")
            
            headers = self.auth.obter_headers_autenticados()
            url = f"https://graph.microsoft.com/v1.0/me/drive/root:{pasta_destino}{nome_arquivo}:/content"
            
            upload_headers = headers.copy()
            upload_headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
            response = requests.put(url, headers=upload_headers, data=excel_bytes, timeout=60)
            
            if response.status_code in [200, 201]:
                logger.info(f"✅ Planilha salva no OneDrive: {pasta_destino}{nome_arquivo}")
                return True
            else:
                logger.error(f"❌ Erro salvando no OneDrive: HTTP {response.status_code}")
                raise Exception(f"Erro HTTP {response.status_code}")
                
        except Exception as e:
            logger.error(f"Erro salvando OneDrive: {e}")
            raise


# Job automático
def job_automatico_06h():
    """Job automático 06:00h"""
    try:
        logger.info("Iniciando job automático 06:00h")
        
        hoje = datetime.now()
        mes = hoje.month
        ano = hoje.year
        
        generator = ExcelGeneratorBRK()
        excel_bytes = generator.gerar_planilha_mensal(mes, ano)
        
        generator._salvar_onedrive_background(excel_bytes, mes, ano)
        
        logger.info(f"Job 06:00h concluído: {mes}/{ano}")
        
    except Exception as e:
        logger.error(f"Erro job_automatico_06h: {e}")
